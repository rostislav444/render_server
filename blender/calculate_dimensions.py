import os
import re

import bpy
import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

exclude_parts = ['fitting', 'legs']

default_prices = {
    'base': {
        'paint': 2400,
    },
    "second": {
        'paint': 2800,
    },
    'frame': {
        'paint': 2800,
    }
}


translation = {
    'legs': 'Ножки',
    'fitting': 'Фурнитура',
    'base': 'Фасады',
    'second': 'Полки',
    'frame': 'Корпус',
}


def get_base_name(name):
    return re.split(r'\.\d+', name)[0]


def get_dimensions_and_copies(obj):
    if obj.name not in bpy.context.view_layer.objects:
        dims = [
            round(abs(obj.dimensions.x) * 1000),
            round(abs(obj.dimensions.y) * 1000),
            round(abs(obj.dimensions.z) * 1000)
        ]
        dims.sort()
        return [dims[1], dims[2], dims[0]], 1

    mirror_count = 1
    array_count = 1

    mirror_mod = None
    for mod in obj.modifiers:
        if mod.type == 'MIRROR':
            mirror_mod = mod
            if any(mod.use_axis):
                mirror_count = 2
        elif mod.type == 'ARRAY':
            array_count *= mod.count

    if mirror_mod:
        was_visible = mirror_mod.show_viewport
        mirror_mod.show_viewport = False
        bpy.context.view_layer.update()

        dims = [
            round(abs(obj.dimensions.x) * 1000),
            round(abs(obj.dimensions.y) * 1000),
            round(abs(obj.dimensions.z) * 1000)
        ]

        mirror_mod.show_viewport = was_visible
        bpy.context.view_layer.update()
    else:
        dims = [
            round(abs(obj.dimensions.x) * 1000),
            round(abs(obj.dimensions.y) * 1000),
            round(abs(obj.dimensions.z) * 1000)
        ]

    dims.sort()
    thickness = dims[0]
    other_dims = sorted(dims[1:], reverse=True)

    dims_list = [other_dims[0], other_dims[1], thickness]
    return dims_list, mirror_count * array_count


def analyze_furniture_parts(length_groups=None, excluded_groups=None):
    if length_groups is None:
        length_groups = ['legs']
    if excluded_groups is None:
        excluded_groups = ['fitting']

    groups = {}
    total_area = 0

    mesh_objects = [obj for obj in bpy.data.objects if obj.type == 'MESH']

    for obj in mesh_objects:
        base_name = get_base_name(obj.name)
        dims_list, copies = get_dimensions_and_copies(obj)

        if base_name not in groups:
            groups[base_name] = {
                'parts': [],
                'area': 0
            }
            if base_name in length_groups:
                groups[base_name]['length'] = 0

        for _ in range(copies):
            groups[base_name]['parts'].append(dims_list)

        if base_name in length_groups:
            length = dims_list[0] / 1000 * copies
            groups[base_name]['length'] = round(groups[base_name]['length'] + length, 2)

        if base_name not in excluded_groups:
            area = (dims_list[0] / 1000) * (dims_list[1] / 1000) * copies
            groups[base_name]['area'] = round(groups[base_name]['area'] + area, 5)
            total_area += area

    groups['total_area'] = round(total_area, 5)
    return groups


def process_blend_files(folder_path):
    # Создаем новый Excel файл
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # Удаляем стандартный лист

    # Перебираем все .blend файлы в папке
    for filename in os.listdir(folder_path):
        if filename.endswith('.blend'):
            blend_path = os.path.join(folder_path, filename)

            # Загружаем .blend файл
            bpy.ops.wm.open_mainfile(filepath=blend_path)

            # Анализируем детали
            data = analyze_furniture_parts(length_groups=['legs'], excluded_groups=['fitting'])

            # Создаем новый лист с именем файла (без расширения)
            sheet_name = os.path.splitext(filename)[0][:31]  # Excel ограничивает длину имени листа
            sheet = wb.create_sheet(sheet_name)

            # Записываем данные на лист
            write_to_excel(data, sheet)

    # Сохраняем Excel файл
    excel_path = os.path.join(folder_path, 'furniture_dimensions.xlsx')
    wb.save(excel_path)


def write_to_excel(data, sheet, start_row=1):
    # Стили
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
    group_font = Font(bold=True)
    total_font = Font(bold=True, color='000066')

    # Заголовки основной таблицы
    headers = ['Ширина', 'Высота', 'Глубина', 'Площадь',
               'Стоимость м²', 'Коэф. покраски', 'Цена покраски м²',
               'Стоимость материала', 'Стоимость покраски']

    # Записываем заголовки
    for col, header in enumerate(headers, 1):
        cell = sheet.cell(row=start_row, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill

    current_row = start_row + 1
    group_totals = {'material': [], 'paint': []}

    # Записываем данные для каждой группы
    for group_name, group_data in data.items():
        if group_name in ['total_area', *exclude_parts]:
            continue

        # Записываем название группы
        group_cell = sheet.cell(row=current_row, column=1)
        group_cell.value = translation.get(group_name, group_name)
        group_cell.font = group_font

        # Устанавливаем цены по умолчанию
        if group_name in default_prices:
            material_price = default_prices[group_name].get('material', 0)
            paint_price = default_prices[group_name].get('paint', 0)
            sheet.cell(row=current_row, column=5, value=material_price)
            sheet.cell(row=current_row, column=7, value=paint_price)
        else:
            sheet.cell(row=current_row, column=5, value=0)
            sheet.cell(row=current_row, column=7, value=0)

        material_price_cell = f'E{current_row}'
        paint_price_cell = f'G{current_row}'

        current_row += 1
        start_sum_row = current_row

        # Записываем размеры деталей
        for part in group_data['parts']:
            # Размеры
            for col, value in enumerate(part, 1):
                sheet.cell(row=current_row, column=col, value=value)

            # Площадь
            area = (part[0] / 1000) * (part[1] / 1000)
            area_cell = sheet.cell(row=current_row, column=4, value=round(area, 6))

            # Коэффициент покраски (по умолчанию 1)
            sheet.cell(row=current_row, column=6, value=1)

            # Формулы расчета стоимости
            material_cost = f'={material_price_cell}*D{current_row}'
            paint_cost = f'={paint_price_cell}*D{current_row}*F{current_row}'

            sheet.cell(row=current_row, column=8, value=material_cost)
            sheet.cell(row=current_row, column=9, value=paint_cost)

            current_row += 1

        # Записываем суммы для группы под ценами
        material_sum = f'=SUM(H{start_sum_row}:H{current_row - 1})'
        paint_sum = f'=SUM(I{start_sum_row}:I{current_row - 1})'

        # Сохраняем ячейки с суммами для итогового подсчета
        group_totals['material'].append(f'E{start_sum_row - 1}*SUM(D{start_sum_row}:D{current_row - 1})')
        group_totals['paint'].append(f'G{start_sum_row - 1}*SUM(D{start_sum_row}:D{current_row - 1})')

        # Добавляем суммы под ценами
        sheet.cell(row=start_sum_row - 1, column=8, value=material_sum).font = group_font
        sheet.cell(row=start_sum_row - 1, column=9, value=paint_sum).font = group_font

        current_row += 1

        # Записываем суммы для группы под ценами
        material_sum = f'=SUM(H{start_sum_row}:H{current_row - 1})'
        paint_sum = f'=SUM(I{start_sum_row}:I{current_row - 1})'

        # Сохраняем ячейки с суммами для итогового подсчета
        group_totals['material'].append(f'E{start_sum_row - 1}*SUM(D{start_sum_row}:D{current_row - 1})')
        group_totals['paint'].append(f'G{start_sum_row - 1}*SUM(D{start_sum_row}:D{current_row - 1})')

        # Добавляем суммы под ценами
        sheet.cell(row=start_sum_row - 1, column=8, value=material_sum).font = group_font
        sheet.cell(row=start_sum_row - 1, column=9, value=paint_sum).font = group_font

        current_row += 1  # Пустая строка между группами

    # Добавляем итоговые суммы
    total_row = current_row
    sheet.cell(row=total_row, column=7, value='ИТОГО:').font = total_font

    # Формула для суммирования значений в столбцах "Стоимость материала" и "Стоимость покраски"
    # Берем только ячейки с суммами групп (которые рядом с ценами)
    material_sums = []
    paint_sums = []

    row = start_row + 1
    while row < total_row:
        cell = sheet.cell(row=row, column=1).value
        if cell and not any(str(x) in str(cell) for x in range(10)):  # Если это название группы
            material_sums.append(f'H{row}')
            paint_sums.append(f'I{row}')
        row += 1

    # Записываем формулы итоговых сумм
    material_total = '+'.join(material_sums)
    paint_total = '+'.join(paint_sums)

    sheet.cell(row=total_row, column=8, value=f'={material_total}').font = total_font
    sheet.cell(row=total_row, column=9, value=f'={paint_total}').font = total_font

    # Общая сумма
    total_row += 1
    sheet.cell(row=total_row, column=7, value='ВСЕГО:').font = total_font
    total_sum = f'=H{total_row - 1}+I{total_row - 1}'
    sheet.cell(row=total_row, column=8, value=total_sum).font = total_font

    # Устанавливаем ширину колонок
    for col in range(1, 10):
        sheet.column_dimensions[get_column_letter(col)].width = 15

    # Настраиваем форматирование для числовых колонок
    money_cols = [5, 7, 8, 9]  # Колонки с ценами
    area_cols = [4]  # Колонки с площадью

    # Задаем форматы для всех строк в колонках
    for row in range(start_row + 1, total_row + 1):
        for col in money_cols:
            cell = sheet.cell(row=row, column=col)
            if isinstance(cell.value, str) and cell.value.startswith('='):
                cell.number_format = '#,##0.00'

        for col in area_cols:
            cell = sheet.cell(row=row, column=col)
            if isinstance(cell.value, (float, int)) or \
                    (isinstance(cell.value, str) and cell.value.startswith('=')):
                cell.number_format = '#,##0.000'


# Получаем путь к папке через input
# folder_path = input("Введите путь к папке с .blend файлами: ")
folder_path = "/Users/rostislavnikolaiev/Desktop/furniture/Charlson/Комоды/h1200w1200"
process_blend_files(folder_path)
